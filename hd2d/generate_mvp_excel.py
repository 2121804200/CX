"""
生成 MVP 数据表 Excel 模板
基于 16_数据表模板.md 规范
包含：角色属性、武器、魔物、技能、Buff 的空白模板 + MVP 初始测试数值
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

# ========== 样式定义 ==========
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBHEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
DATA_FONT = Font(name="微软雅黑", size=10)
NOTE_FONT = Font(name="微软雅黑", size=9, italic=True, color="808080")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
DATA_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淡黄：示例数据行

def set_cell(ws, row, col, value, font=DATA_FONT, fill=DATA_FILL, alignment=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.border = THIN_BORDER
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    return cell

def write_header_row(ws, row, headers, col_start=1):
    """写入表头行（深蓝底白字）"""
    for i, h in enumerate(headers):
        set_cell(ws, row, col_start + i, h, font=HEADER_FONT, fill=HEADER_FILL,
                 alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

def write_data_row(ws, row, data, col_start=1, is_example=True):
    """写入数据行，示例数据用淡黄底"""
    fill = EXAMPLE_FILL if is_example else DATA_FILL
    for i, val in enumerate(data):
        set_cell(ws, row, col_start + i, val, fill=fill,
                 alignment=Alignment(horizontal="center", vertical="center"))

def write_note_row(ws, row, note, col_count):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    set_cell(ws, row, 1, note, font=NOTE_FONT, fill=DATA_FILL)

def auto_width(ws, min_width=8, max_width=36):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    # 中文字符按2宽度算
                    char_len = 0
                    for ch in line:
                        char_len += 2 if ord(ch) > 127 else 1
                    max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))

# ========== 创建工作簿 ==========
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "说明"

# ========== 说明页 ==========
ws.merge_cells("A1:G1")
set_cell(ws, 1, 1, "星轨转生：异世界魔女纪事 — MVP 数据表", font=TITLE_FONT, fill=DATA_FILL)
ws.merge_cells("A2:G2")
set_cell(ws, 2, 1, f"生成日期：{datetime.now().strftime('%Y-%m-%d')}  版本：V1.0-MVP", font=Font(name="微软雅黑", size=10, color="666666"), fill=DATA_FILL)

notes = [
    "",
    "📋 本文件包含以下数据表（每个Sheet一张表）：",
    "  ① 角色属性  — character_attr.csv  — 主角各阶位基础属性",
    "  ② 武器防具  — weapon.csv         — 新手装备模板（剑/杖/袍/帽/指环）",
    "  ③ 魔物      — monster.csv        — 初始区域魔物（史莱姆/狼/蝙蝠）",
    "  ④ 技能      — skill.csv          — 一阶法术 + 武器奥义 + 魔物技能",
    "  ⑤ Buff     — buff.csv           — 基础增益/减益状态",
    "",
    "🎨 配色说明：",
    "  - 深蓝底白字 = 字段名行",
    "  - 浅蓝底     = 字段类型/说明行",
    "  - 淡黄底     = MVP 示例数据（正式配表时替换为实际数值）",
    "  - 行号后带 ★ = MVP 填充了测试数值的行",
    "",
    "⚙️ 所有表格 UTF-8 编码，逗号分隔，首行为字段名。",
    "⚙️ 黄色底色的行为示例数据，正式配表时可清空或替换。",
]
for i, note in enumerate(notes):
    ws.merge_cells(start_row=4 + i, start_column=1, end_row=4 + i, end_column=7)
    set_cell(ws, 4 + i, 1, note, font=Font(name="微软雅黑", size=10), fill=DATA_FILL)

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 80

# =====================================================================
# Sheet 1: 角色属性表
# =====================================================================
ws1 = wb.create_sheet("①角色属性")
char_headers = [
    "id", "name", "rank", "hp_max", "mp_max", "atk", "def", "matk", "mdef", "spd",
    "fire_apt", "water_apt", "wind_apt", "earth_apt", "heal_apt", "dark_apt", "time_apt",
    "sword_mastery", "alchemy_mastery", "breakthrough_cost", "note"
]
char_types = [
    "string", "string", "int(1~7)", "int", "int", "int", "int", "int", "int", "int",
    "int(0~100)", "int(0~100)", "int(0~100)", "int(0~100)", "int(0~100)", "int(0~100)", "int(0~100)",
    "int", "int", "string", "string"
]

write_header_row(ws1, 1, char_headers)
write_note_row(ws1, 2, "类型/说明：" + " | ".join(char_types), len(char_headers))

# ★ MVP 数据：玩家初阶
char_data = ["player_lv1", "转生魔女·初阶", 1, 200, 100, 12, 8, 22, 12, 100,
     60, 50, 55, 45, 40, 30, 10, 0, 0, "POTION_BREAK_1", "MVP初始属性"]
write_data_row(ws1, 3, char_data)
write_note_row(ws1, 4, "★ 上面这行是 MVP 测试数值：开局主角 1 阶基础面板，魔法倾向，元素适性为后续法术学习留空间", len(char_headers))

# 空行提示
write_note_row(ws1, 6, "以下留空，供后续阶位扩展填写（rank 2~7）", len(char_headers))
for r in range(7, 13):
    empty_row = [""] * len(char_headers)
    empty_row[0] = f"player_lv{r-6}"  # rank 2~7 id 提示
    write_data_row(ws1, r, empty_row, is_example=False)

auto_width(ws1)

# =====================================================================
# Sheet 2: 武器防具表
# =====================================================================
ws2 = wb.create_sheet("②武器防具")
weapon_headers = [
    "id", "name", "type", "weapon_type", "rank_require", "quality",
    "atk", "def", "matk", "mdef", "hp_bonus", "mp_bonus", "spd_bonus",
    "skill_id", "enchant_max", "durability_max", "source", "recipe_id", "sell_price", "desc"
]
weapon_types = [
    "string", "string", "int(1武器2头盔3护甲4饰品)", "int(1剑士2魔法3暗黑,仅武器填)", "int", "int(1白~6红)",
    "int", "int", "int", "int", "int", "int", "int",
    "string", "int", "int", "int(1掉落2合成3商店4宝箱5任务)", "string", "int", "string"
]

write_header_row(ws2, 1, weapon_headers)
write_note_row(ws2, 2, "类型/说明：" + " | ".join(weapon_types), len(weapon_headers))

# ★ MVP 新手装备数据
weapon_data = [
    ["W_NOVICE_SWORD", "新手铁剑", 1, 1, 1, 1,     8, 0, 0, 0, 0, 0, 0,  "SKILL_SLASH_01", 1, 80,  1, "", 30,  "新手剑士的铁剑，附带横斩奥义。"],
    ["W_NOVICE_STAFF", "新手魔杖", 1, 2, 1, 1,     0, 0, 10, 0, 0, 15, 0, "",               1, 70,  3, "", 35,  "魔女学徒的木制法杖，略微提升魔力。"],
    ["W_NOVICE_HAT",   "新手魔女帽", 2, 0, 1, 1,   0, 2, 0, 5, 10, 0, 0,  "",               1, 60,  3, "", 20,  "宽檐魔女帽，提供少量魔法防护。"],
    ["W_NOVICE_ROBE",  "新手魔女袍", 3, 0, 1, 1,   0, 5, 0, 2, 20, 0, 0,  "",               1, 90,  3, "", 25,  "粗布缝制的魔女长袍，轻便但防护有限。"],
    ["W_NOVICE_RING",  "新手指环",   4, 0, 1, 1,   0, 0, 3, 1, 0, 5, 0,   "",               1, 50,  3, "", 15,  "镶嵌小颗魔力石的指环。"],
]
for r, d in enumerate(weapon_data):
    write_data_row(ws2, 3 + r, d)

write_note_row(ws2, 8, "★ 上面 5 行是 MVP 新手五件套：剑（物理）+ 杖（魔法）+ 帽 + 袍 + 指环，rank_require=1，品质=1(白)", len(weapon_headers))
write_note_row(ws2, 10, "以下留空，供后续品质/阶位装备填写", len(weapon_headers))
auto_width(ws2)

# =====================================================================
# Sheet 3: 魔物表
# =====================================================================
ws3 = wb.create_sheet("③魔物")
monster_headers = [
    "id", "name", "monster_type", "rank", "hp_max", "mp_max",
    "atk", "def", "matk", "mdef", "spd",
    "element_main", "res_fire", "res_water", "res_wind", "res_earth", "res_heal", "res_dark", "res_time",
    "skill_group_id", "passive_buff_id", "drop_group_id",
    "exp_gain", "mastery_gain_sword", "mastery_gain_magic",
    "is_cycle_boss", "control_resist", "desc"
]
monster_types = [
    "string", "string", "int(1普通2精英3BOSS)", "int(1~7)", "int", "int",
    "int", "int", "int", "int", "int",
    "int(1火2水3风4土5治愈6暗黑7时空)", "int(%)", "int(%)", "int(%)", "int(%)", "int(%)", "int(%)", "int(%)",
    "string", "string", "string",
    "int", "int", "int",
    "int(0否1是)", "float(0~1)", "string"
]

write_header_row(ws3, 1, monster_headers)
write_note_row(ws3, 2, "类型/说明：" + " | ".join(monster_types), len(monster_headers))

# ★ MVP 初始魔物数据（rank 1 新手区）
monster_data = [
    ["MOB_SLIME_01", "毒棘史莱姆", 1, 1, 80, 30,    8, 4, 6, 8, 60,   4, -10, 0, -20, 20, 0, 0, 0,  "SKG_SLIME", "", "DROP_SLIME_01",   5, 1, 1, 0, 0.0, "森林中常见的史莱姆，体表带毒刺。弱风。"],
    ["MOB_WOLF_01",  "森林狼",     1, 1, 110, 20,  14, 6, 0, 3, 90,   0,  0,   0,  0,   0,  0, 0, 0,  "SKG_WOLF",  "", "DROP_WOLF_01",    8, 2, 0, 0, 0.0, "敏捷的森林掠食者，群居。速度快。"],
    ["MOB_BAT_01",   "暗影蝙蝠",   1, 1, 60, 40,   5, 2, 12, 5, 75,   6,  0,   0,  0,   0,  0, 10, 0, "SKG_BAT",   "", "DROP_BAT_01",     6, 0, 2, 0, 0.0, "洞穴中的暗属性蝙蝠，会使用超声波攻击。"],
]
for r, d in enumerate(monster_data):
    write_data_row(ws3, 3 + r, d)

write_note_row(ws3, 6, "★ 上面 3 行是 MVP 初始魔物：史莱姆(土弱风)、森林狼(高速物理)、暗影蝙蝠(暗属性法系)，均为 rank=1 普通怪", len(monster_headers))
write_note_row(ws3, 8, "以下留空，供精英/BOSS 魔物扩展填写", len(monster_headers))
auto_width(ws3)

# =====================================================================
# Sheet 4: 技能表
# =====================================================================
ws4 = wb.create_sheet("④技能")
skill_headers = [
    "id", "name", "skill_type", "element", "target_type",
    "mp_cost", "hp_cost", "base_damage", "damage_type",
    "atk_coef", "matk_coef", "cast_time", "cooldown",
    "apply_buff_id", "apply_buff_chance", "rank_require", "is_ultimate",
    "effect_prefab", "desc"
]
skill_types = [
    "string", "string", "int(1普攻2元素魔法3武器奥义4魔物技能5被动)", "int(0无1火2水3风4土5治愈6暗黑7时空)", "int(1单敌2全敌3单友4全友5自身)",
    "int", "int", "int(治疗为正)", "int(1物理2魔法3真实)",
    "float", "float", "float(秒)", "int(回合)",
    "string", "float(0~1)", "int", "int(0否1是)",
    "string", "string"
]

write_header_row(ws4, 1, skill_headers)
write_note_row(ws4, 2, "类型/说明：" + " | ".join(skill_types), len(skill_headers))

# ★ MVP 一阶法术 + 武器奥义 + 魔物技能
skill_data = [
    # 主角法术（一阶元素魔法，rank 1）
    ["SKILL_FIRE_BALL",   "火球术",   2, 1, 1, 15, 0, 40, 2, 0.0, 1.2, 0.5, 0, "",           0.0, 1, 0, "fx_fireball",   "发射一颗火球，造成火属性魔法伤害。"],
    ["SKILL_WATER_ARROW", "水箭术",   2, 2, 1, 12, 0, 35, 2, 0.0, 1.0, 0.4, 0, "",           0.0, 1, 0, "fx_waterarrow", "凝聚水元素射出箭矢，造成水属性魔法伤害。"],
    ["SKILL_WIND_BLADE",  "风刃术",   2, 3, 1, 10, 0, 30, 2, 0.0, 0.9, 0.3, 0, "",           0.0, 1, 0, "fx_windblade",  "挥出无形风刃，速度快消耗低。"],
    ["SKILL_EARTH_SPIKE", "地刺术",   2, 4, 1, 14, 0, 38, 2, 0.0, 1.1, 0.6, 0, "",           0.0, 1, 0, "fx_earthspike", "从地面刺出岩石，造成土属性魔法伤害。"],
    ["SKILL_HEAL_LIGHT",  "小治愈术", 2, 5, 3, 10, 0, 35, 2, 0.0, 0.8, 0.5, 0, "",           0.0, 1, 0, "fx_heal",       "以光元素治愈单体友方少量生命值。"],
    # 武器奥义（绑定新手铁剑）
    ["SKILL_SLASH_01",    "横斩",     3, 0, 1, 5,  0, 25, 1, 1.2, 0.0, 0.2, 1, "",           0.0, 1, 1, "fx_slash",      "新手铁剑的奥义横斩，造成物理伤害。"],
    # 魔物技能
    ["SKILL_MOB_POISON",  "毒液喷射", 4, 4, 1, 5,  0, 15, 2, 0.0, 0.5, 0.5, 0, "BUFF_POISON",   0.5, 0, 0, "fx_poison",     "史莱姆喷射毒液，概率附加中毒状态。"],
    ["SKILL_MOB_CLAW",    "爪击",     4, 0, 1, 0,  0, 18, 1, 1.0, 0.0, 0.2, 0, "",           0.0, 0, 0, "fx_claw",       "森林狼的利爪攻击。"],
    ["SKILL_MOB_ECHO",    "超声波",   4, 6, 1, 8,  0, 20, 2, 0.0, 0.7, 0.4, 0, "BUFF_SLOW",      0.3, 0, 0, "fx_echo",       "暗影蝙蝠发出超声波，概率使目标减速。"],
]
for r, d in enumerate(skill_data):
    write_data_row(ws4, 3 + r, d)

write_note_row(ws4, 12, "★ 上面 9 行 MVP 技能：火球/水箭/风刃/地刺/治愈 5 个一阶法术 + 横斩奥义 + 3 个魔物技能", len(skill_headers))
write_note_row(ws4, 13, "  技能类型：2=元素魔法，3=武器奥义，4=魔物技能", len(skill_headers))
write_note_row(ws4, 15, "以下留空，供高阶法术、更多奥义、被动技能扩展", len(skill_headers))
auto_width(ws4)

# =====================================================================
# Sheet 5: Buff表
# =====================================================================
ws5 = wb.create_sheet("⑤Buff")
buff_headers = [
    "id", "name", "buff_type", "effect_type", "effect_value", "is_percent",
    "duration", "max_stack", "refresh_on_apply", "can_dispel", "is_debuff",
    "icon", "tick_every_turn", "desc"
]
buff_types = [
    "string", "string", "int(1增益2减益)", "int(1加攻2加防3加速4加魔攻5灼烧6中毒7减速8眩晕9沉默10护盾11吸血12阶位压制)",
    "int", "int(0固定1%)", "int(回合)", "int", "int(0否1是)", "int(0否1是)", "int(0否1是)",
    "string", "int(0否1是)", "string"
]

write_header_row(ws5, 1, buff_headers)
write_note_row(ws5, 2, "类型/说明：" + " | ".join(buff_types), len(buff_headers))

# ★ MVP 基础 Buff
buff_data = [
    ["BUFF_ATK_UP",   "攻击提升", 1, 1,  20, 1, 3, 1, 1, 1, 0, "icon_buff_atk",    0, "物理攻击力提升20%，持续3回合。"],
    ["BUFF_DEF_UP",   "防御提升", 1, 2,  20, 1, 3, 1, 1, 1, 0, "icon_buff_def",    0, "物理防御力提升20%，持续3回合。"],
    ["BUFF_MATK_UP",  "魔攻提升", 1, 4,  20, 1, 3, 1, 1, 1, 0, "icon_buff_matk",   0, "魔法攻击力提升20%，持续3回合。"],
    ["BUFF_SPD_UP",   "加速",     1, 3,  30, 1, 2, 1, 1, 1, 0, "icon_buff_spd",    0, "速度提升30%，持续2回合。"],
    ["BUFF_POISON",   "中毒",     2, 6,  5,  1, 3, 3, 1, 1, 1, "icon_debuff_poison", 1, "每回合损失最大生命值5%，持续3回合，最多叠加3层。"],
    ["BUFF_SHOCK_BURN","灼烧",    2, 5,  3,  1, 3, 1, 1, 1, 1, "icon_debuff_burn",   1, "每回合损失最大生命值3%，持续3回合。"],
    ["BUFF_SLOW",     "减速",     2, 7,  30, 1, 2, 1, 0, 1, 1, "icon_debuff_slow",   0, "速度降低30%，持续2回合。"],
    ["BUFF_STUN",     "眩晕",     2, 8,  0,  0, 1, 1, 0, 1, 1, "icon_debuff_stun",   0, "无法行动1回合。"],
    ["BUFF_SHIELD",   "魔力护盾", 1, 10, 30, 0, 3, 1, 0, 1, 0, "icon_buff_shield",  0, "获得可吸收30点伤害的护盾，持续3回合。"],
]
for r, d in enumerate(buff_data):
    write_data_row(ws5, 3 + r, d)

write_note_row(ws5, 12, "★ 上面 9 行 MVP Buff：4 增益（攻/防/魔攻/速）+ 4 减益（中毒/灼烧/减速/眩晕）+ 1 护盾", len(buff_headers))
write_note_row(ws5, 13, "  中毒、灼烧 tick_every_turn=1 表示每回合自动结算扣血；中毒 max_stack=3 可叠加至 15%/tick", len(buff_headers))
write_note_row(ws5, 15, "以下留空，供更多 Buff/Debuff 扩展（如沉默、阶位压制、吸血等）", len(buff_headers))
auto_width(ws5)

# ========== 保存 ==========
output_path = r"d:\untiy\hd2d\MVP数据表_初始测试数值.xlsx"
wb.save(output_path)
print(f"✅ 已生成：{output_path}")
print(f"   Sheet 数：{len(wb.sheetnames)}")
for name in wb.sheetnames:
    print(f"   - {name}")
